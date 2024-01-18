import os
import sys
import optparse
import time
import json
import openpyxl
import datetime
from google.cloud import monitoring_v3
# from google.cloud.monitoring_v3.types.common import TypedValue


def extractDatabaseName(instanceId):
    return instanceId.split("/")[-1]

def get_command_by_args(commands, *args):
    count = 0
    for cmd in args:
        try:
            count += commands[cmd]
        except KeyError:
            pass
    return count

def get_all_commands(commands):
    count = 0
    for value in commands.values():
        count += value
    return count

def processNodeStats(processedMetricPoints):
    nodeStats = {
        'Throughput (Ops)': 0, 
        'GetTypeCmds': 0, 
        'SetTypeCmds': 0, 
        'OtherTypeCmds': 0, 
        'BitmapBasedCmds': 0, 
        'ClusterBasedCmds': 0, 
        'EvalBasedCmds': 0, 
        'GeoSpatialBasedCmds': 0, 
        'HashBasedCmds': 0, 
        'HyperLogLogBasedCmds': 0, 
        'KeyBasedCmds': 0, 
        'ListBasedCmds': 0, 
        'PubSubBasedCmds': 0, 
        'SetBasedCmds': 0, 
        'SortedSetBasedCmds': 0, 
        'StringBasedCmds': 0, 
        'StreamBasedCmds': 0, 
        'TransactionBasedCmds': 0
    }

    for processedMetricPoint in processedMetricPoints.values():
        for key, value in processedMetricPoint.items():
            if value > nodeStats[key]:
                nodeStats[key] = value

    return nodeStats

def processMetricPoint(metricPoint):
    processedMetricPoint = {}
    
    processedMetricPoint['Throughput (Ops)'] = round(get_all_commands(metricPoint))

    # Get type commands
    processedMetricPoint['GetTypeCmds'] = round(get_command_by_args(
        metricPoint, 
        'bitcount',
        'bitfield_ro',
        'bitpos',
        'getbit',
        'geodist',
        'geohash',
        'geopos',
        'georadiusbymember_ro',
        'georadius_ro',
        'geosearch',
        'hexists',
        'hget',
        'hgetall',
        'hkeys',
        'hlen',
        'hmget',
        'hrandfield',
        'hscan',
        'hstrlen',
        'hvals',
        'pfcount', 
        'dump',
        'exists',
        'expiretime',
        'keys',
        'pexpiretime',
        'pttl',
        'randomkey',
        'scan',
        'sort',
        'sort_ro',
        'touch',
        'ttl',
        'type',
        'lindex',
        'llen',
        'lpos',
        'lrange',
        'scard',
        'sdiff',
        'sinter',
        'sintercard',
        'sismember',
        'smembers',
        'smismember',
        'srandmember',
        'sscan',
        'sunion',
        'zcard',
        'zcount',
        'zdiff',
        'zinter',
        'zintercard',
        'zlexcount',
        'zmscore',
        'zrandmember',
        'zrange',
        'zrangebylex',
        'zrangebyscore',
        'zrank',
        'zrevrange',
        'zrevrank',
        'zscan',
        'zscore',
        'zunion',
        'get',
        'getrange',
        'lcs',
        'mget',
        'strlen',
        'substr',
        'xinfo',
        'xlen',
        'xpending',
        'xrange',
        'xread',
        'xrevrange'
    ))

    # Set type commands
    processedMetricPoint['SetTypeCmds'] = round(get_command_by_args(
        metricPoint, 
        'bitfield',
        'bitop',
        'setbit',
        'geoadd',
        'georadius',
        'georadiusbymember',
        'geosearchstore',
        'hdel',
        'hincrby',
        'hincrbyfloat',
        'hmset',
        'hset',
        'hsetnx',
        'pfadd',
        'pfdebug',
        'pfmerge',
        'copy',
        'del',
        'expire',
        'expireat',
        'migrate',
        'move',
        'persist',
        'pexpire',
        'pexpireat',
        'rename',
        'renamenx',
        'restore',
        'sort',
        'unlink',
        'blmove',
        'blmpop',
        'blpop',
        'brpop',
        'brpoplpush',
        'linsert',
        'lmove',
        'lmpop',
        'lpop',
        'lpush',
        'lpushx',
        'lrem',
        'lset',
        'ltrim',
        'rpop',
        'rpoplpush',
        'rpush',
        'rpushx',
        'sadd',
        'sdiffstore',
        'sinterstore',
        'smove',
        'spop',
        'srem',
        'sunionstore',
        'bzmpop',
        'bzpopmax',
        'bzpopmin',
        'zadd',
        'zdiffstore',
        'zincrby',
        'zinterstore',
        'zmpop',
        'zpopmax',
        'zpopmin',
        'zrangestore',
        'zrem',
        'zremrangebylex',
        'zremrangebyrank',
        'zremrangebyscore',
        'zrevrangebylex',
        'zrevrangebyscore',
        'zunionstore',
        'append',
        'decr',
        'decrby',
        'getdel',
        'getex',
        'getset',
        'incr',
        'incrby',
        'incrbyfloat',
        'mset',
        'msetnx',
        'psetex',
        'set',
        'setex',
        'setnx',
        'setrange',
        'xack',
        'xadd',
        'xautoclaim',
        'xclaim',
        'xdel',
        'xgroup',
        'xreadgroup',
        'xsetid',
        'xtrim'
    ))

    # Other type commands
    processedMetricPoint['OtherTypeCmds'] = round(get_command_by_args(
        metricPoint, 
        'asking',
        'cluster',
        'readonly',
        'readwrite',
        'auth',
        'client',
        'echo',
        'hello',
        'ping',
        'quit',
        'reset',
        'select',
        'eval',
        'evalsha',
        'evalsha_ro',
        'eval_ro',
        'fcall',
        'fcall_ro',
        'function',
        'script',
        'pfselftest',
        'object',
        'wait',
        'psubscribe',
        'publish',
        'pubsub',
        'punsubscribe',
        'spublish',
        'ssubscribe',
        'subscribe',
        'sunsubscribe',
        'unsubscribe',
        'discard',
        'exec',
        'multi',
        'unwatch',
        'watch'
    ))

    # Bitmaps based commands
    processedMetricPoint['BitmapBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'bitcount',
        'bitfield',
        'bitfield_ro',
        'bitop',
        'bitpos',
        'getbit',
        'setbit'
    ))

    # Cluster based commands
    processedMetricPoint['ClusterBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'asking',
        'cluster',
        'readonly',
        'readwrite'
    ))

    # Eval based commands
    processedMetricPoint['EvalBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'eval',
        'evalsha',
        'evalsha_ro',
        'eval_ro',
        'fcall',
        'fcall_ro',
        'function',
        'script'
    ))

    # GeoSpatial based commands
    processedMetricPoint['GeoSpatialBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'geoadd',
        'geodist',
        'geohash',
        'geopos',
        'georadius',
        'georadiusbymember',
        'georadiusbymember_ro',
        'georadius_ro',
        'geosearch',
        'geosearchstore'
    ))

    # Hash based commands
    processedMetricPoint['HashBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'hdel',
        'hexists',
        'hget',
        'hgetall',
        'hincrby',
        'hincrbyfloat',
        'hkeys',
        'hlen',
        'hmget',
        'hmset',
        'hrandfield',
        'hscan',
        'hset',
        'hsetnx',
        'hstrlen',
        'hvals'
    ))

    # HyperLogLog based commands
    processedMetricPoint['HyperLogLogBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'pfadd',
        'pfcount',
        'pfdebug',
        'pfmerge',
        'pfselftest'
    ))

    # Keys based commands
    processedMetricPoint['KeyBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'copy',
        'del',
        'dump',
        'exists',
        'expire',
        'expireat',
        'expiretime',
        'keys',
        'migrate',
        'move',
        'object',
        'persist',
        'pexpire',
        'pexpireat',
        'pexpiretime',
        'pttl',
        'randomkey',
        'rename',
        'renamenx',
        'restore',
        'scan',
        'sort',
        'sort_ro',
        'touch',
        'ttl',
        'type',
        'unlink',
        'wait'
    ))

    # List based commands
    processedMetricPoint['ListBasedCmds'] = round(get_command_by_args(
        metricPoint,
        'blmove',
        'blmpop',
        'blpop',
        'brpop',
        'brpoplpush',
        'lindex',
        'linsert',
        'llen',
        'lmove',
        'lmpop',
        'lpop',
        'lpos',
        'lpush',
        'lpushx',
        'lrange',
        'lrem',
        'lset',
        'ltrim',
        'rpop',
        'rpoplpush',
        'rpush',
        'rpushx'    
    ))

    # PubSub based commands
    processedMetricPoint['PubSubBasedCmds'] = round(get_command_by_args(
        metricPoint,
        'psubscribe',
        'publish',
        'pubsub',
        'punsubscribe',
        'spublish',
        'ssubscribe',
        'subscribe',
        'sunsubscribe',
        'unsubscribe'
    ))

    # Sets based commands
    processedMetricPoint['SetBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'sadd',
        'scard',
        'sdiff',
        'sdiffstore',
        'sinter',
        'sintercard',
        'sinterstore',
        'sismember',
        'smembers',
        'smismember',
        'smove',
        'spop',
        'srandmember',
        'srem',
        'sscan',
        'sunion',
        'sunionstore'
    ))

    # SortedSets based commands
    processedMetricPoint['SortedSetBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'bzmpop',
        'bzpopmax',
        'bzpopmin',
        'zadd',
        'zcard',
        'zcount',
        'zdiff',
        'zdiffstore',
        'zincrby',
        'zinter',
        'zintercard',
        'zinterstore',
        'zlexcount',
        'zmpop',
        'zmscore',
        'zpopmax',
        'zpopmin',
        'zrandmember',
        'zrange',
        'zrangebylex',
        'zrangebyscore',
        'zrangestore',
        'zrank',
        'zrem',
        'zremrangebylex',
        'zremrangebyrank',
        'zremrangebyscore',
        'zrevrange',
        'zrevrangebylex',
        'zrevrangebyscore',
        'zrevrank',
        'zscan',
        'zscore',
        'zunion',
        'zunionstore'
    ))

    # String based commands
    processedMetricPoint['StringBasedCmds'] = round(get_command_by_args(
        metricPoint, 
        'append',
        'decr',
        'decrby',
        'get',
        'getdel',
        'getex',
        'getrange',
        'getset',
        'incr',
        'incrby',
        'incrbyfloat',
        'lcs',
        'mget',
        'mset',
        'msetnx',
        'psetex',
        'set',
        'setex',
        'setnx',
        'setrange',
        'strlen',
        'substr'
    ))

    # Stream based commands
    processedMetricPoint['StreamBasedCmds'] = round(get_command_by_args(
        metricPoint,
        'xack',
        'xadd',
        'xautoclaim',
        'xclaim',
        'xdel',
        'xgroup',
        'xinfo',
        'xlen',
        'xpending',
        'xrange',
        'xread',
        'xreadgroup',
        'xrevrange',
        'xsetid',
        'xtrim'
    ))
    
    # Transaction based commands
    processedMetricPoint['TransactionBasedCmds'] = round(get_command_by_args(
        metricPoint,
        'discard',
        'exec',
        'multi',
        'unwatch',
        'watch'
    ))
    
    return processedMetricPoint

def process_google_service_account(service_account, projectID):

    if projectID:
        project_id = projectID
    else: 
        try:
            f = open (service_account, "r")
            data = json.loads(f.read())
            f.close()
            project_id = data['project_id']
            if not project_id:
                raise Exception("Invalid json file")
        except:
            return

    # Set the value GOOGLE_APPLICATION_CREDENTIALS variable
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = service_account
    print("Processing Google Account with credentials found in: ", os.environ['GOOGLE_APPLICATION_CREDENTIALS'])

    client = monitoring_v3.MetricServiceClient()
    project_name = f"projects/{project_id}"

    now = time.time()
    seconds = int(now)
    nanos = int((now - seconds) * 10 ** 9)

    interval = monitoring_v3.TimeInterval(
        {
            "end_time": {"seconds": seconds, "nanos": nanos},
            "start_time": {"seconds": (seconds - 604800), "nanos": nanos},
        }
    )
    aggregation = monitoring_v3.Aggregation(
        {
            "alignment_period": {"seconds": 60},
            "per_series_aligner": monitoring_v3.Aggregation.Aligner.ALIGN_RATE,
            "cross_series_reducer": monitoring_v3.Aggregation.Reducer.REDUCE_NONE
        }
    )

    metric_points = {}

    # Check the following resources for more metrics
    # https://cloud.google.com/memorystore/docs/redis/supported-monitoring-metrics


    # Call the google cloud "redis.googleapis.com/commands/calls" to get commandstats
    results = client.list_time_series(
        request={
            "name": project_name,
            "filter": 'metric.type = "redis.googleapis.com/commands/calls"',
            "interval": interval,
            "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
            "aggregation": aggregation,
        }
    )

    for result in results:
        database = extractDatabaseName(result.resource.labels["instance_id"])
        node_id = result.resource.labels["node_id"]
        cmd = result.metric.labels['cmd']

        if not database in metric_points:
            metric_points[database] = {}
        if not node_id in metric_points[database]:
            metric_points[database][node_id] = { 
                "Source": "MS",
                "ClusterId": project_id + '.' + database,
                "InstanceId": result.resource.labels["instance_id"],
                "NodeId": node_id,
                "NodeRole": "Master" if result.metric.labels['role'] == 'primary' else "Replica",
                "NodeType": "",
                "Region": result.resource.labels["region"],
                "points": {}
            }
        
        for point in result.points:
            interval = point.interval.start_time.timestamp()
            if not interval in metric_points[database][node_id]["points"]:
                metric_points[database][node_id]["points"][interval] = {}
            
            point_value = 0
            if (result.value_type == 2):
                point_value = point.value.int64_value
            elif (result.value_type == 3):
                point_value = point.value.double_value

            metric_points[database][node_id]["points"][interval][cmd] = point_value

    for database in metric_points:
        for node in metric_points[database]:
            if not "processed_points" in metric_points[database][node]:
                metric_points[database][node]["processed_points"] = {}
            for point in metric_points[database][node]["points"]:
                metric_points[database][node]["processed_points"][point] = processMetricPoint(metric_points[database][node]["points"][point])

    for database in metric_points:
        for node in metric_points[database]:
            metric_points[database][node]['commandstats'] = processNodeStats(metric_points[database][node]["processed_points"])
            del metric_points[database][node]["points"]
            del metric_points[database][node]["processed_points"] 

    # CurrItems	
    #Call the google cloud metrics "redis.googleapis.com/stats/memory/usage" to get "BytesUsedForCache"
    interval = monitoring_v3.TimeInterval()
    now = time.time()
    seconds = int(now)
    nanos = int((now - seconds) * 10 ** 9)
    interval = monitoring_v3.TimeInterval(
        {
            "end_time": {"seconds": seconds, "nanos": nanos},
            "start_time": {"seconds": (seconds - 604800), "nanos": nanos},
        }
    )

    # Retrieve MaxMemory (a.k.a. Capacity)
    results = client.list_time_series(
        request={
            "name": project_name,
            "filter": 'metric.type = "redis.googleapis.com/stats/memory/maxmemory"',
            "interval": interval,
            "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
        }
    ) 
    for result in results:
        database = extractDatabaseName(result.resource.labels["instance_id"])
        node_id = result.resource.labels["node_id"]
        MaxMemory = 0
        for point in result.points:
            if point.value.int64_value > MaxMemory:
                MaxMemory = point.value.int64_value
        if database in metric_points:
           metric_points[database][node_id]['MaxMemory'] = MaxMemory


    results = client.list_time_series(
        request={
            "name": project_name,
            "filter": 'metric.type = "redis.googleapis.com/stats/memory/usage"',
            "interval": interval,
            "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
        }
    )
    for result in results:
        database = extractDatabaseName(result.resource.labels["instance_id"])
        node_id = result.resource.labels["node_id"]
        BytesUsedForCache = 0
        for point in result.points:
            if point.value.int64_value > BytesUsedForCache:
                BytesUsedForCache = point.value.int64_value
        if database in metric_points:
           metric_points[database][node_id]['BytesUsedForCache'] = BytesUsedForCache


    # CacheHits	
    # CacheMisses	
    #TODO: Call the google cloud metrics "redis.googleapis.com/clients/connected" to get "CurrConnections"
    # NetworkBytesIn	
    # NetworkBytesOut	
    # NetworkPacketsIn	
    # NetworkPacketsOut	
    # EngineCPUUtilization	
    #TODO: Call the google cloud metrics "redis.googleapis.com/stats/evicted_keys" to get "Evictions"
    # ReplicationBytes	
    # ReplicationLag

    return project_id, metric_points

def create_workbooks(outDir, projects, report_name):

    # Open an Excel workbook
    output_file_path = "%s/%s.xlsx" % (outDir, report_name) 
    try:
       # Try to open an existing workbook
       wb = openpyxl.load_workbook(output_file_path)
       print("Workbook loaded successfully.")
    except FileNotFoundError:
       # If the file doesn't exist, create a new workbook
       wb = openpyxl.Workbook()
       print("New workbook created.")    
    ws = wb.active
    ws.title = 'ClusterData'

    for project in projects:
        for cluster in projects[project]:
            for node in projects[project][cluster]:
                
                node_commandstats = projects[project][cluster][node]['commandstats']                
                node_info = projects[project][cluster][node]
                del node_info['commandstats']
                node_stats = {**node_info, **node_commandstats}

                if node_stats is not None:
                    if ws.max_row == 1:
                        ws.append(list(node_stats.keys()))    
                    ws.append(list(node_stats.values()))
    
    print(f"Writing output file {output_file_path}")
    wb.save(output_file_path)

def main():
    if not sys.version_info >= (3, 6):
        print("Please upgrade python to a version at least 3.6")
        exit(1)

    parser = optparse.OptionParser()
    parser.add_option(
        "-d", 
        "--out-dir", 
        dest="outDir", 
        default=".",
        help="The directory to output the results. If the directory does not exist the script will try to create it.", 
        metavar="PATH"
    )
    parser.add_option(
        "-r",
        "--report-name",
        dest="report_name",
        default="ms-report-" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S"),
        help="The name of the msstats report.",
        metavar="REPORT_NAME"
    )  
    parser.add_option(
        "-p",
        "--project-id",
        dest="project_id",
        default="",
        help="The Google Cloud Project ID containing MemoryStore instances.",
        metavar="PROJECT_ID"
    )

    (options, _) = parser.parse_args()

    if not os.path.isdir(options.outDir):
        os.makedirs(options.outDir)

    # Scan for .json files in order to find the service account files
    path_to_json = '.'
    service_accounts = [os.path.abspath(os.path.join(path_to_json, pos_json)) for pos_json in os.listdir(path_to_json) if pos_json.endswith('.json')]

    projects = {}
    # For each service account found try to fetch the clusters metrics using the 
    # google cloud monitoring api metrics
    for service_account in service_accounts:
        project_id, stats = process_google_service_account(service_account, options.project_id)
        projects[project_id] = stats

    create_workbooks(options.outDir, projects, options.report_name)

    print("Done!")


if __name__ == "__main__":
    main()
