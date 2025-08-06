import unittest
import os
import tempfile
import json
from unittest.mock import patch, MagicMock
import openpyxl

from msstats import (
    extractDatabaseName,
    get_command_by_args,
    get_all_commands,
    processNodeStats,
    processMetricPoint,
    get_project_from_service_account_and_authenticate,
    process_google_project,
    create_workbooks,
)


class TestMSSStats(unittest.TestCase):
    """Test suite for msstats utility functions"""

    def test_extractDatabaseName(self):
        """Test extracting database name from instance ID"""
        # Test typical instance ID format
        instance_id = (
            "projects/test-project/locations/us-central1/instances/redis-instance"
        )
        expected = "redis-instance"
        result = extractDatabaseName(instance_id)
        self.assertEqual(result, expected)

        # Test with different format
        instance_id = "projects/my-project/locations/europe-west1/instances/my-redis-db"
        expected = "my-redis-db"
        result = extractDatabaseName(instance_id)
        self.assertEqual(result, expected)

        # Test edge case with single part
        instance_id = "redis-simple"
        expected = "redis-simple"
        result = extractDatabaseName(instance_id)
        self.assertEqual(result, expected)

    def test_get_command_by_args(self):
        """Test getting command counts by specific arguments"""
        commands = {"get": 100, "set": 50, "del": 25, "incr": 10, "exists": 5}

        # Test single command
        result = get_command_by_args(commands, "get")
        self.assertEqual(result, 100)

        # Test multiple commands
        result = get_command_by_args(commands, "get", "set")
        self.assertEqual(result, 150)

        # Test with non-existent command
        result = get_command_by_args(commands, "nonexistent")
        self.assertEqual(result, 0)

        # Test mix of existing and non-existent commands
        result = get_command_by_args(commands, "get", "nonexistent", "set")
        self.assertEqual(result, 150)

        # Test empty args
        result = get_command_by_args(commands)
        self.assertEqual(result, 0)

    def test_get_all_commands(self):
        """Test getting total count of all commands"""
        commands = {"get": 100, "set": 50, "del": 25, "incr": 10}

        result = get_all_commands(commands)
        self.assertEqual(result, 185)

        # Test empty commands dict
        result = get_all_commands({})
        self.assertEqual(result, 0)

        # Test single command
        result = get_all_commands({"test": 42})
        self.assertEqual(result, 42)

    def test_processNodeStats(self):
        """Test processing node statistics to get maximum values"""
        # Create sample processed metric points from multiple nodes
        processed_metric_points = {
            "node1": {
                "Throughput (Ops)": 100,
                "GetTypeCmds": 50,
                "SetTypeCmds": 30,
                "OtherTypeCmds": 20,
                "BitmapBasedCmds": 0,
                "ClusterBasedCmds": 0,
                "EvalBasedCmds": 0,
                "GeoSpatialBasedCmds": 0,
                "HashBasedCmds": 10,
                "HyperLogLogBasedCmds": 0,
                "KeyBasedCmds": 5,
                "ListBasedCmds": 0,
                "PubSubBasedCmds": 0,
                "SetBasedCmds": 0,
                "SortedSetBasedCmds": 0,
                "StringBasedCmds": 15,
                "StreamBasedCmds": 0,
                "TransactionBasedCmds": 0,
            },
            "node2": {
                "Throughput (Ops)": 150,  # Higher than node1
                "GetTypeCmds": 40,
                "SetTypeCmds": 60,  # Higher than node1
                "OtherTypeCmds": 10,
                "BitmapBasedCmds": 5,  # Higher than node1
                "ClusterBasedCmds": 0,
                "EvalBasedCmds": 0,
                "GeoSpatialBasedCmds": 0,
                "HashBasedCmds": 8,
                "HyperLogLogBasedCmds": 0,
                "KeyBasedCmds": 12,  # Higher than node1
                "ListBasedCmds": 0,
                "PubSubBasedCmds": 0,
                "SetBasedCmds": 0,
                "SortedSetBasedCmds": 0,
                "StringBasedCmds": 10,
                "StreamBasedCmds": 0,
                "TransactionBasedCmds": 0,
            },
        }

        result = processNodeStats(processed_metric_points)

        # Should take maximum values across all nodes
        self.assertEqual(result["Throughput (Ops)"], 150)  # max from node2
        self.assertEqual(result["GetTypeCmds"], 50)  # max from node1
        self.assertEqual(result["SetTypeCmds"], 60)  # max from node2
        self.assertEqual(result["BitmapBasedCmds"], 5)  # max from node2
        self.assertEqual(result["KeyBasedCmds"], 12)  # max from node2

    def test_processMetricPoint(self):
        """Test processing a single metric point"""
        # Create sample metric point data
        metric_point = {
            "get": 50,
            "set": 30,
            "hget": 20,  # Hash command
            "lpush": 10,  # List command
            "sadd": 5,  # Set command
            "unknown_cmd": 15,  # Should be counted in total but not categorized
        }

        result = processMetricPoint(metric_point)

        # Check that all commands are counted in throughput
        expected_throughput = sum(metric_point.values())
        self.assertEqual(result["Throughput (Ops)"], expected_throughput)

        # Check that hash commands are properly categorized
        self.assertEqual(result["HashBasedCmds"], 20)  # hget

        # Check that list commands are properly categorized
        self.assertEqual(result["ListBasedCmds"], 10)  # lpush

        # Check that set commands are properly categorized
        self.assertEqual(result["SetBasedCmds"], 5)  # sadd

        # Verify structure - all expected keys should be present
        expected_keys = {
            "Throughput (Ops)",
            "GetTypeCmds",
            "SetTypeCmds",
            "OtherTypeCmds",
            "BitmapBasedCmds",
            "ClusterBasedCmds",
            "EvalBasedCmds",
            "GeoSpatialBasedCmds",
            "HashBasedCmds",
            "HyperLogLogBasedCmds",
            "KeyBasedCmds",
            "ListBasedCmds",
            "PubSubBasedCmds",
            "SetBasedCmds",
            "SortedSetBasedCmds",
            "StringBasedCmds",
            "StreamBasedCmds",
            "TransactionBasedCmds",
        }
        self.assertEqual(set(result.keys()), expected_keys)

    def test_processMetricPoint_empty(self):
        """Test processing empty metric point"""
        result = processMetricPoint({})

        # All values should be 0
        for value in result.values():
            self.assertEqual(value, 0)

    def test_processNodeStats_single_node(self):
        """Test processing node stats with single node"""
        processed_metric_points = {
            "node1": {
                "Throughput (Ops)": 100,
                "GetTypeCmds": 50,
                "SetTypeCmds": 30,
                "OtherTypeCmds": 20,
                "BitmapBasedCmds": 0,
                "ClusterBasedCmds": 0,
                "EvalBasedCmds": 0,
                "GeoSpatialBasedCmds": 0,
                "HashBasedCmds": 10,
                "HyperLogLogBasedCmds": 0,
                "KeyBasedCmds": 5,
                "ListBasedCmds": 0,
                "PubSubBasedCmds": 0,
                "SetBasedCmds": 0,
                "SortedSetBasedCmds": 0,
                "StringBasedCmds": 15,
                "StreamBasedCmds": 0,
                "TransactionBasedCmds": 0,
            }
        }

        result = processNodeStats(processed_metric_points)

        # Should return the same values as the single node
        self.assertEqual(result["Throughput (Ops)"], 100)
        self.assertEqual(result["GetTypeCmds"], 50)
        self.assertEqual(result["HashBasedCmds"], 10)

    def test_processNodeStats_empty(self):
        """Test processing empty node stats"""
        result = processNodeStats({})

        # Should return default structure with all zeros
        expected_keys = {
            "Throughput (Ops)",
            "GetTypeCmds",
            "SetTypeCmds",
            "OtherTypeCmds",
            "BitmapBasedCmds",
            "ClusterBasedCmds",
            "EvalBasedCmds",
            "GeoSpatialBasedCmds",
            "HashBasedCmds",
            "HyperLogLogBasedCmds",
            "KeyBasedCmds",
            "ListBasedCmds",
            "PubSubBasedCmds",
            "SetBasedCmds",
            "SortedSetBasedCmds",
            "StringBasedCmds",
            "StreamBasedCmds",
            "TransactionBasedCmds",
        }

        self.assertEqual(set(result.keys()), expected_keys)
        for value in result.values():
            self.assertEqual(value, 0)


class TestMSStatsIntegration(unittest.TestCase):
    """Integration tests for msstats functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.test_project_id = "test-project-123"

        # Create a mock service account file with required fields
        self.service_account_data = {
            "project_id": self.test_project_id,
            "type": "service_account",
            "private_key_id": "test-key-id",
            "private_key": "Some Mock Private Key\n",
            "client_email": "test@test-project.iam.gserviceaccount.com",
            "client_id": "123456789",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
        }

        self.service_account_file = os.path.join(
            self.temp_dir, "test-service-account.json"
        )
        with open(self.service_account_file, "w") as f:
            json.dump(self.service_account_data, f)

    def tearDown(self):
        """Clean up test fixtures"""
        if os.path.exists(self.service_account_file):
            os.remove(self.service_account_file)
        os.rmdir(self.temp_dir)

    @patch("msstats.monitoring_v3.MetricServiceClient")
    def test_process_google_project_with_mock_data(self, mock_client_class):
        """Test processing service account with mocked Google Cloud responses"""
        # Mock the monitoring client
        mock_client = MagicMock()
        mock_client_class.return_value = mock_client

        # Mock time series response for commands/calls
        mock_result = MagicMock()
        mock_result.resource.labels = MagicMock()
        mock_result.resource.labels.__getitem__ = MagicMock(
            side_effect=lambda key: {
                "instance_id": "projects/test-project/locations/us-central1/instances/test-redis",
                "region": "us-central1",
                "node_id": "node-0",
            }[key]
        )

        mock_result.metric.labels = MagicMock()
        mock_result.metric.labels.__getitem__ = MagicMock(
            side_effect=lambda key: {"cmd": "get", "role": "primary"}[key]
        )

        # Mock data points
        mock_point = MagicMock()
        mock_point.interval.start_time.timestamp.return_value = (
            1640995200.0  # 2022-01-01
        )
        mock_point.value.int64_value = 100
        mock_result.points = [mock_point]
        mock_result.value_type = 2  # INT64

        # Mock memory usage response
        mock_memory_result = MagicMock()
        mock_memory_result.resource.labels = MagicMock()
        mock_memory_result.resource.labels.__getitem__ = MagicMock(
            side_effect=lambda key: {
                "instance_id": "projects/test-project/locations/us-central1/instances/test-redis",
                "node_id": "node-0",
            }[key]
        )
        mock_memory_point = MagicMock()
        mock_memory_point.value.int64_value = 1000000  # 1MB
        mock_memory_result.points = [mock_memory_point]

        # Set up client responses
        mock_client.list_time_series.side_effect = [
            [mock_result],  # commands/calls response
            [mock_memory_result],  # memory/usage response
            [mock_memory_result],  # memory/maxmemory response
        ]

        # Test the function
        stats = process_google_project(
            self.test_project_id,
            duration=3600,  # 1 hour
            step=60,
        )

        # Assertions
        self.assertIsInstance(stats, dict)
        self.assertIn("test-redis", stats)

        # Check that the database stats contain expected structure
        database_stats = stats["test-redis"]
        self.assertIn("node-0", database_stats)

        node_stats = database_stats["node-0"]
        self.assertIn("commandstats", node_stats)
        self.assertIn("Source", node_stats)
        self.assertIn("ClusterId", node_stats)
        self.assertIn("NodeRole", node_stats)
        self.assertEqual(node_stats["NodeRole"], "Master")

    def test_create_workbooks_integration(self):
        """Test creating Excel workbooks from processed data"""
        # Create sample processed data
        projects_data = {
            "test-project": {
                "redis-instance-1": {
                    "node-0": {
                        "Source": "MS",
                        "ClusterId": "redis-instance-1",
                        "NodeId": "node-0",
                        "NodeRole": "Master",
                        "NodeType": "",
                        "Region": "us-central1",
                        "Project ID": "test-project",
                        "InstanceId": "projects/test-project/locations/us-central1/instances/redis-instance-1",
                        "BytesUsedForCache": 1000000,
                        "MaxMemory": 2000000,
                        "commandstats": {
                            "Throughput (Ops)": 500,
                            "GetTypeCmds": 300,
                            "SetTypeCmds": 200,
                            "OtherTypeCmds": 0,
                            "BitmapBasedCmds": 0,
                            "ClusterBasedCmds": 0,
                            "EvalBasedCmds": 0,
                            "GeoSpatialBasedCmds": 0,
                            "HashBasedCmds": 0,
                            "HyperLogLogBasedCmds": 0,
                            "KeyBasedCmds": 0,
                            "ListBasedCmds": 0,
                            "PubSubBasedCmds": 0,
                            "SetBasedCmds": 0,
                            "SortedSetBasedCmds": 0,
                            "StringBasedCmds": 0,
                            "StreamBasedCmds": 0,
                            "TransactionBasedCmds": 0,
                        },
                    }
                }
            }
        }

        # Create workbooks
        create_workbooks(self.temp_dir, projects_data)

        # Verify Excel file was created
        excel_file = os.path.join(self.temp_dir, "test-project.xlsx")
        self.assertTrue(os.path.exists(excel_file))

        # Verify Excel file contents
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        self.assertEqual(ws.title, "ClusterData")

        # Check that we have headers and data
        self.assertGreater(ws.max_row, 1)  # Should have header + at least 1 data row
        self.assertGreater(ws.max_column, 10)  # Should have many columns

        # Check some expected headers are present
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Source", headers)
        self.assertIn("ClusterId", headers)
        self.assertIn("NodeRole", headers)
        self.assertIn("Throughput (Ops)", headers)

        # Clean up
        wb.close()
        os.remove(excel_file)

    @patch("msstats.monitoring_v3.MetricServiceClient")
    def test_service_account_file_parsing(self, mock_client_class):
        """Test that service account files are parsed correctly"""
        # Mock the monitoring client
        mock_client = MagicMock()
        mock_client_class.return_value = mock_client
        mock_client.list_time_series.return_value = []

        # It should read from file:
        project_id = get_project_from_service_account_and_authenticate(
            self.service_account_file
        )
        self.assertEqual(project_id, self.test_project_id)

    def test_invalid_service_account_file(self):
        """Test handling of invalid service account files"""
        # Create invalid JSON file
        invalid_file = os.path.join(self.temp_dir, "invalid.json")
        with open(invalid_file, "w") as f:
            f.write("invalid json content")

        # Should return None for invalid files
        result = get_project_from_service_account_and_authenticate(invalid_file)
        self.assertIsNone(result)

        # Clean up
        os.remove(invalid_file)

    def test_missing_service_account_file(self):
        """Test handling of missing service account files"""
        nonexistent_file = os.path.join(self.temp_dir, "nonexistent.json")

        # Should return None for missing files
        result = get_project_from_service_account_and_authenticate(nonexistent_file)
        self.assertIsNone(result)


if __name__ == "__main__":
    unittest.main()
